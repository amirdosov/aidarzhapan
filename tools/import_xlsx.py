# -*- coding: utf-8 -*-
"""
Импорт родословной из source/АйдарЖапан ШЕЖІРЕСІ-2.xlsx в data/people.js.

Как читается файл (подробности — в docs/DATA.md):

  Лист2 — визуальная схема. Каждое поколение на своей строке, заливка
          ячейки кодирует пол и статус. Отсюда берётся состав людей.
  Лист1 — та же родословная вложенными ячейками: родитель — объединённая
          ячейка, растянутая на строки всех потомков, колонка = поколение.
          Отсюда берутся связи и те люди, кого на схеме не поместилось
          (ветки Нияза и Кенжебая).

  Для колен 2-15 родитель однозначен: это человек прямой линии из
  предыдущего колена (в файле она пронумерована и залита маджентой).

  Строка на Лист2 — это возрастная волна, а не колено: дети Олжабая
  1955-1969 годов стоят в одной строке с внуками Қосбалы. Поэтому колено
  считается по цепочке от Адая, а строка остаётся подсказкой и попадает
  в отчёт как row_gen.

  Всё, что автоматика не берёт, лежит в tools/manual.json и накладывается
  сверху. Расхождения печатаются в tools/report.txt.

Запуск:  python tools/import_xlsx.py
"""

import difflib
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "source" / "АйдарЖапан ШЕЖІРЕСІ-2.xlsx"
MANUAL = ROOT / "tools" / "manual.json"
OUT_JS = ROOT / "data" / "people.js"
REPORT = ROOT / "tools" / "report.txt"

# --- разметка листов ------------------------------------------------------

# Лист2: строка -> номер волны (примерно колено)
GEN_ROWS = {10: 1, 12: 2, 14: 3, 16: 4, 20: 5, 22: 6, 25: 7, 29: 8,
            31: 9, 33: 10, 35: 11, 38: 12, 40: 13, 42: 14, 44: 15,
            46: 16, 48: 17, 50: 18, 52: 19, 54: 20}

# Лист2: строки-подписи (Мақпал бәйбішеден, 3-ші әйелінен) — это не люди
NOTE_ROWS = {18, 23, 27, 36}
# Отдельные ячейки на строках людей, где на самом деле пояснение
NOTE_CELLS = {"U31"}

# Лист1: колонка -> номер волны; строка 8 — шапка с годами
GEN_COLS = {"N": 15, "O": 16, "P": 17, "Q": 18, "R": 19, "S": 20}
FIRST_ROW = 9

# Лист1: ячейки, которые не человек, а пояснение к родителю
L1_NOTES = {"O22": "6 қыз бала", "O24": "5 қыздың аты жөні белгісіз",
            "O25": None}
# Лист1: в одной ячейке несколько имён через запятую
L1_SPLIT = {"Q22"}

FILL_MAIN = "FFFF66FF"    # маджента — прямая линия
FILL_MALE = "FFFF99FF"    # розовый — мужчины рода
FILL_OTHER = "FFFFFF00"   # жёлтый — женщины и жиендер

# --- нормализация имён ----------------------------------------------------

FOLD = str.maketrans({
    "ә": "а", "ғ": "г", "қ": "к", "ң": "н", "ө": "о", "ұ": "у",
    "ү": "у", "һ": "х", "і": "и", "ё": "е", "й": "и",
})

# слова-пометки, которые именем не являются
MARKS = {"ж", "жыл", "апа", "апаи", "ага", "жиен", "кыз", "бала", "балалары",
         "иа", "или", "аты", "белгисиз", "жони", "деген", "ба"}
STOP = MARKS | {"ер", "т", "б", "не", "г", "р", "атамыз", "анамыз"}

YEAR = r"(?<!\d)(?:1[3-9]\d{2}|20[0-3]\d)(?!\d)"


def fold(s):
    """Свести казахскую и русскую орфографию к общему виду для сравнения."""
    s = unicodedata.normalize("NFC", s or "").lower().translate(FOLD)
    return re.sub(r"[^a-zа-я0-9]+", " ", s).strip()


def tokens(name):
    return [t for t in fold(name).split() if t and t not in STOP and not t.isdigit()]


TRANSLIT = {
    "а": "a", "ә": "a", "б": "b", "в": "v", "г": "g", "ғ": "g", "д": "d",
    "е": "e", "ё": "e", "ж": "zh", "з": "z", "и": "i", "й": "i", "к": "k",
    "қ": "q", "л": "l", "м": "m", "н": "n", "ң": "n", "о": "o", "ө": "o",
    "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ұ": "u", "ү": "u",
    "ф": "f", "х": "h", "һ": "h", "ц": "c", "ч": "ch", "ш": "sh",
    "щ": "sch", "ъ": "", "ы": "y", "і": "i", "ь": "", "э": "e",
    "ю": "yu", "я": "ya",
}


def slug(name):
    out = "".join(
        TRANSLIT.get(ch, ch if (ch.isascii() and ch.isalnum()) else "-")
        for ch in unicodedata.normalize("NFC", name).lower()
    )
    return re.sub(r"-+", "-", out).strip("-") or "x"


def strip_marks(name):
    """Убрать из имени пометки: Бауыржан жиен -> Бауыржан."""
    keep = [w for w in name.split() if fold(w) not in MARKS]
    out = " ".join(keep).strip(" .-")
    return out if len(out) >= 3 else name


# --- разбор текста ячейки -------------------------------------------------

def dedupe(text):
    """В объединённых ячейках одна и та же подпись повторена по разу
    на каждую накрытую строку. Оставляем первую."""
    parts = [x.strip() for x in re.split(r"\s{3,}", str(text)) if x.strip()]
    if len(parts) > 1:
        first = fold(parts[0])
        if all(difflib.SequenceMatcher(None, first, fold(x)).ratio() > 0.8
               for x in parts[1:]):
            return parts[0]
    return str(text)


def parse_cell(text):
    """Из текста ячейки достать имя, годы, второе имя и пометки."""
    raw = re.sub(r"\s+", " ", dedupe(text)).strip()
    s = re.sub(r"^\d{1,2}\s*[.)]\s*", "", raw)   # номер прямой линии: 17. Адай

    born = died = None
    m = re.search(YEAR + r"\s*[-–—]\s*(?:" + YEAR + r"|\d{2,3}\?)", s)
    if m:
        a, b = re.split(r"\s*[-–—]\s*", m.group(0), maxsplit=1)
        born, died = int(a), (int(b) if b.isdigit() else None)
        s = s.replace(m.group(0), " ")
    else:
        m = re.search(YEAR, s)
        if m:
            born = int(m.group(0))
            s = s.replace(m.group(0), " ", 1)

    flags = set()
    low = fold(raw)
    if "кыз бала" in low or re.search(r"\bапаи?\b", low):
        flags.add("f")
    if "жиен" in low:
        flags.add("zhien")
    if "егиз" in low:
        flags.add("twin")
    if "кенже бала" in low:
        flags.add("kenzhe")

    # скобки: либо второе имя (Рустам), либо пометка (қыз бала)
    alt = None
    for inner in re.findall(r"\(([^)]*)\)", s):
        if tokens(inner):
            alt = strip_marks(inner.strip())
    s = re.sub(r"\([^)]*\)", " ", s)

    # хвостовая проза после запятой или слэша — в примечание
    note = None
    parts = re.split(r"[,/]", s, maxsplit=1)
    if len(parts) == 2 and len(parts[1].strip(" .-?")) > 3:
        note = parts[1].strip(" .-")
    s = parts[0]

    s = re.sub(r"\bж(ыл)?\b\.?", " ", s)
    s = re.sub(r"[.\-–—:;?]+", " ", s)
    s = re.sub(r"^\d+\s+", "", s)
    s = re.sub(r"\s+", " ", s).strip()

    # если после чистки осталось предложение, а не имя — уводим в примечание
    words = s.split()
    if len(words) > 3:
        note = ((note + ". ") if note else "") + " ".join(words[1:])
        s = words[0]

    # примечание из одних пометок (қыз бала) ничего не добавляет
    if note and not [t for t in fold(note).split()
                     if t not in MARKS and not t.isdigit()]:
        note = None

    return {"name": strip_marks(s), "alt": alt, "born": born, "died": died,
            "flags": flags, "note": note, "raw": raw}


def sex_of(flags, color=None):
    if color == FILL_OTHER:
        return "m" if "zhien" in flags else "f"
    if color in (FILL_MAIN, FILL_MALE):
        return "m"
    return "f" if "f" in flags else None


def fill_of(cell):
    try:
        rgb = cell.fill.fgColor.rgb
        return rgb if isinstance(rgb, str) else None
    except Exception:
        return None


# --- Лист2: состав людей --------------------------------------------------

def read_roster(ws):
    people = []
    for row, gen in sorted(GEN_ROWS.items()):
        for cell in ws[row]:
            if cell.value is None or cell.coordinate in NOTE_CELLS:
                continue
            color = fill_of(cell)
            if color not in (FILL_MAIN, FILL_MALE, FILL_OTHER):
                continue  # без заливки — рассказ или подпись, не человек
            p = parse_cell(cell.value)
            if not p["name"]:
                continue
            p.update(row_gen=gen, cell=cell.coordinate, sheet="Лист2",
                     main=color == FILL_MAIN, sex=sex_of(p["flags"], color))
            p["line"] = ("main" if p["main"] else
                         "zhien" if "zhien" in p["flags"] else
                         "rod" if color == FILL_MALE else "out")
            people.append(p)
    return people


# --- Лист1: связи ---------------------------------------------------------

def read_blocks(ws):
    """Ячейки колонок N..S как блоки строк: колонка, начало, конец, текст."""
    spans, ghosts = {}, []
    for rng in ws.merged_cells.ranges:
        cols = [get_column_letter(c) for c in range(rng.min_col, rng.max_col + 1)]
        if cols[0] not in GEN_COLS:
            continue
        spans[(cols[0], rng.min_row)] = rng.max_row
        # ячейка растянута на несколько колонок (Олжабай O124:P137) —
        # в следующих колонках оставляем тень, чтобы внуки нашли родителя
        for extra in cols[1:]:
            if extra in GEN_COLS:
                ghosts.append((extra, rng.min_row, rng.max_row, cols[0]))

    blocks, index = [], {}
    for col, gen in GEN_COLS.items():
        seen = set()
        for cell in ws[col]:
            if cell.row < FIRST_ROW or cell.value is None:
                continue
            text = re.sub(r"\s+", " ", dedupe(cell.value)).strip()
            if not text or not re.search(r"[А-Яа-яӘҚҢӨҰҮҺІ]", text):
                continue  # пустое или сплошные вопросительные знаки
            seen.add(cell.row)
            b = {"col": col, "gen": gen, "r0": cell.row,
                 "r1": spans.get((col, cell.row), cell.row),
                 "text": text, "coord": cell.coordinate, "ghost": False}
            blocks.append(b)
            index[cell.coordinate] = b
        # объединённая ячейка без текста — тоже человек: автор не вписал себя
        for (c, r0), r1 in spans.items():
            if c == col and r0 not in seen and r0 >= FIRST_ROW:
                b = {"col": col, "gen": gen, "r0": r0, "r1": r1, "text": "",
                     "coord": col + str(r0), "ghost": False}
                blocks.append(b)
                index[b["coord"]] = b

    for col, r0, r1, origin_col in ghosts:
        blocks.append({"col": col, "gen": GEN_COLS[col], "r0": r0, "r1": r1,
                       "text": "", "coord": col + str(r0), "ghost": True,
                       "origin": origin_col + str(r0)})

    blocks.sort(key=lambda b: (b["gen"], b["r0"]))
    for b in blocks:
        if b["ghost"]:
            b["origin_block"] = index.get(b["origin"])
    return blocks


def link_blocks(blocks):
    """Родитель блока — блок предыдущей колонки, накрывающий его строки."""
    by_gen = defaultdict(list)
    for b in blocks:
        by_gen[b["gen"]].append(b)
    for b in blocks:
        b["parent_block"] = None
        for cand in by_gen.get(b["gen"] - 1, []):
            if cand["r0"] <= b["r0"] <= cand["r1"]:
                b["parent_block"] = cand.get("origin_block") or cand
                break
    return blocks


def match_blocks(blocks, roster):
    """Сопоставить блоки Лист1 с людьми Лист2: имя, год, похожесть."""
    by_token = defaultdict(list)
    by_gen = defaultdict(list)
    for p in roster:
        by_gen[p["row_gen"]].append(p)
        for t in tokens(p["name"]) + tokens(p["alt"] or ""):
            by_token[(p["row_gen"], t)].append(p)

    for b in blocks:
        b["person"] = None
        b["parsed"] = parse_cell(b["text"]) if b["text"] and not b["ghost"] else None
        info = b["parsed"]
        if not info or not info["name"]:
            continue

        # 1. по слову имени внутри той же волны
        cands = []
        for t in tokens(info["name"]) + tokens(info["alt"] or ""):
            cands += by_token.get((b["gen"], t), [])
        cands = list({id(c): c for c in cands}.values())

        # тёзок разводим годом: Мансұр 2018 — это Мансур без года, не 2017-го
        if len(cands) > 1 and info["born"]:
            same = [c for c in cands if c["born"] == info["born"]]
            blank = [c for c in cands if not c["born"]]
            cands = same or blank or cands

        # 2. похожесть написания: Жамиля и Жамилә, Кенжебай и Кенжебек
        if not cands:
            scored = [(difflib.SequenceMatcher(None, fold(info["name"]),
                                               fold(p["name"])).ratio(), p)
                      for p in by_gen[b["gen"]]]
            cands = [p for r, p in scored if r >= 0.80]

        # 3. последний ключ — год рождения
        if not cands and info["born"]:
            cands = [p for p in by_gen[b["gen"]] if p["born"] == info["born"]]

        if len(cands) == 1:
            b["person"] = cands[0]
            cands[0]["l1cell"] = b["coord"]
    return blocks


# --- сборка ---------------------------------------------------------------

def adopt(blocks, roster):
    """Люди, которые есть только на Лист1 (ветки Нияза и Кенжебая)."""
    added = []
    for b in blocks:
        if b["person"] or not b.get("parsed") or b["coord"] in L1_NOTES:
            continue
        info = b["parsed"]
        if not info["name"]:
            continue
        names = ([n.strip() for n in info["raw"].split(",") if n.strip()]
                 if b["coord"] in L1_SPLIT else [info["name"]])
        for i, name in enumerate(names):
            p = dict(info, name=strip_marks(name), row_gen=b["gen"],
                     cell=b["coord"], sheet="Лист1", main=False,
                     sex=sex_of(info["flags"]),
                     line="zhien" if "zhien" in info["flags"] else "rod")
            if len(names) > 1:
                p["born"] = p["died"] = p["note"] = None
            roster.append(p)
            added.append(p)
            if i == 0:
                b["person"] = p
            else:
                b.setdefault("also", []).append(p)
    return added


def assign_ids(roster):
    """Идентификаторы выдаём один раз: у кого уже есть — не трогаем."""
    used = defaultdict(int)
    for p in roster:
        if p.get("id"):
            used[slug(p["name"])] += 1
    for p in roster:
        if p.get("id"):
            continue
        base = slug(p["name"])
        used[base] += 1
        p["id"] = base if used[base] == 1 else "%s-%s" % (
            base, p["born"] or used[base])
    return roster


def compute_gen(roster, by_id):
    """Колено считаем от Адая по цепочке родителей, а не по строке файла."""
    memo = {}

    def depth(pid, seen=()):
        if pid in memo:
            return memo[pid]
        p = by_id.get(pid)
        if not p:
            return None
        par = p.get("parent")
        if not par or par in seen:
            # корень — только Адай; остальные без родителя остаются без колена
            value = 1 if (not par and p.get("row_gen") == 1) else None
        else:
            up = depth(par, seen + (pid,))
            value = up + 1 if up else None
        memo[pid] = value
        return value

    for p in roster:
        p["gen"] = depth(p["id"])
    return roster


def main():
    wb = openpyxl.load_workbook(XLSX)
    s1, s2 = wb["Лист1"], wb["Лист2"]

    roster = assign_ids(read_roster(s2))
    blocks = match_blocks(link_blocks(read_blocks(s1)), roster)
    by_id = {p["id"]: p for p in roster}

    manual = json.loads(MANUAL.read_text("utf-8")) if MANUAL.exists() else {}

    # ручная привязка ячеек Лист1: безымянных (Q68 — сам составитель)
    # и тех, где написание разошлось сильнее, чем ловит автоматика
    for coord, pid in manual.get("blocks", {}).items():
        for b in blocks:
            if b["coord"] == coord and pid in by_id:
                b["person"] = by_id[pid]

    added = adopt(blocks, roster)
    assign_ids(roster)
    by_id = {p["id"]: p for p in roster}

    # 1. связи из вложенности Лист1
    for b in blocks:
        pb = b["parent_block"]
        if not b["person"] or not pb or not pb.get("person"):
            continue
        for p in [b["person"]] + b.get("also", []):
            p["parent"] = pb["person"]["id"]
            p["src"] = "Лист1 " + b["coord"]
        # Лист1 знает годы и подробности, которых нет на схеме
        if b.get("parsed") and b["person"].get("sheet") == "Лист2":
            for k in ("born", "died", "note"):
                if b["parsed"].get(k) and not b["person"].get(k):
                    b["person"][k] = b["parsed"][k]

    # 2. колена 2-15: родитель — прямая линия предыдущего колена
    by_row = defaultdict(list)
    for p in roster:
        by_row[p["row_gen"]].append(p)
    for gen in range(2, 16):
        parents = [p for p in by_row.get(gen - 1, []) if p["main"]]
        if len(parents) != 1:
            continue
        for p in by_row.get(gen, []):
            if not p.get("parent"):
                p["parent"] = parents[0]["id"]
                p["src"] = "прямая линия"

    # 3. ручная разметка
    for pid, patch in manual.get("people", {}).items():
        if pid in by_id:
            by_id[pid].update(patch)
            by_id[pid]["src"] = "manual.json"
    for extra in manual.get("extra", []):
        extra.setdefault("src", "manual.json")
        extra.setdefault("sheet", "manual.json")
        extra.setdefault("row_gen", 0)
        roster.append(extra)
        by_id[extra["id"]] = extra
    for child, parent in manual.get("parents", {}).items():
        if child in by_id:
            by_id[child]["parent"] = parent
            by_id[child]["src"] = "manual.json"

    compute_gen(roster, by_id)
    unions = manual.get("unions", [])
    write_report(roster, blocks, by_id, added, unions)
    write_js(roster, unions)
    print("людей: %d | со связью: %d | колено посчитано: %d | отчёт: %s" % (
        len(roster),
        sum(1 for p in roster if p.get("parent")),
        sum(1 for p in roster if p.get("gen")),
        REPORT.relative_to(ROOT)))


def hint(p, by_coord):
    """Подсказка для ячейки Лист1, съехавшей на строку: кто стоит выше."""
    b = by_coord.get(p.get("l1cell") or p.get("cell", ""))
    if not b:
        return ""
    prev = [x for x in by_coord.values()
            if x["gen"] == b["gen"] - 1 and x["r1"] < b["r0"]
            and x.get("person") and b["r0"] - x["r1"] <= 6]
    if not prev:
        return ""
    near = max(prev, key=lambda x: x["r1"])
    return "выше стоит %s (%s)" % (near["person"]["name"], near["coord"])


def write_report(roster, blocks, by_id, added, unions):
    L = ["Отчёт импорта. Сверять глазами, а не на веру.",
         "Пересобрать: python tools/import_xlsx.py", ""]

    by_gen = defaultdict(list)
    for p in roster:
        by_gen[p.get("gen")].append(p)
    L.append("СОСТАВ ПО КОЛЕНАМ (колено от Адая, посчитано по цепочке)")
    for gen in sorted(by_gen, key=lambda g: (g is None, g)):
        names = ", ".join(p["name"] + ("*" if p.get("main") else "")
                          for p in by_gen[gen])
        L.append("  %s (%d): %s" % (
            ("%2d" % gen) if gen else "??", len(by_gen[gen]), names))
    L.append("  * — прямая линия")
    L.append("")

    by_coord = {b["coord"]: b for b in blocks if not b["ghost"]}
    orphans = [p for p in roster if not p.get("parent") and p["name"] != "Адай"]
    L.append("БЕЗ РОДИТЕЛЯ — НУЖНО СПРОСИТЬ У СЕМЬИ: %d" % len(orphans))
    for p in sorted(orphans, key=lambda x: (x["row_gen"], x["name"])):
        L.append("  строка-волна %2d  %-16s %-6s %-6s %-32s %s" % (
            p["row_gen"], p["name"], p.get("born") or "",
            p.get("cell", ""), p.get("raw", "")[:32], hint(p, by_coord)))
    L.append("")

    L.append("ДОБАВЛЕНЫ С ЛИСТ1 (на схеме Лист2 их нет): %d" % len(added))
    for p in added:
        L.append("  %-6s %-16s %-6s пол: %-4s %s" % (
            p["cell"], p["name"], p.get("born") or "", p.get("sex") or "?",
            p.get("parent") or "родитель не найден"))
    L.append("")

    nosex = [p for p in roster if not p.get("sex")]
    L.append("ПОЛ НЕ ОПРЕДЕЛЁН: %d" % len(nosex))
    for p in nosex:
        L.append("  %-16s %s %s" % (p["name"], p.get("cell", ""), p.get("sheet", "")))
    L.append("")

    guessed = [p for p in roster if p.get("sex_src")]
    L.append("ПОЛ ПРОСТАВЛЕН ВРУЧНУЮ — ПОДТВЕРДИТЬ: %d" % len(guessed))
    for p in guessed:
        L.append("  %-16s %-4s (%s)" % (p["name"], p["sex"], p["sex_src"]))
    L.append("")

    shifted = [p for p in roster
               if p.get("gen") and p["row_gen"] and p["gen"] != p["row_gen"]]
    L.append("КОЛЕНО НЕ СОВПАЛО СО СТРОКОЙ ФАЙЛА: %d" % len(shifted))
    for p in shifted:
        L.append("  %-16s строка-волна %2d, по цепочке %2d (через %s)" % (
            p["name"], p["row_gen"], p["gen"], p.get("parent")))
    L.append("")

    unmatched = [b for b in blocks
                 if b.get("parsed") and not b["person"] and b["coord"] not in L1_NOTES]
    L.append("ЛИСТ1: ЯЧЕЙКИ БЕЗ ЧЕЛОВЕКА: %d" % len(unmatched))
    for b in unmatched:
        L.append("  %-5s волна %2d  %s" % (b["coord"], b["gen"], b["text"][:70]))
    L.append("")

    empty = [b for b in blocks if not b["text"] and not b["ghost"]]
    L.append("ЛИСТ1: ПУСТЫЕ ОБЪЕДИНЁННЫЕ ЯЧЕЙКИ: %d" % len(empty))
    for b in empty:
        kids = [x["parsed"]["name"] for x in blocks
                if x.get("parent_block") is b and x.get("parsed")]
        L.append("  %-5s волна %2d  строки %d-%d  дети: %s  %s" % (
            b["coord"], b["gen"], b["r0"], b["r1"], ", ".join(kids) or "-",
            "-> " + b["person"]["name"] if b.get("person") else "НЕ ОПОЗНАН"))
    L.append("")

    bad = ["%s -> %s" % (p["id"], p["parent"]) for p in roster
           if p.get("parent") and p["parent"] not in by_id]
    bad += [" + ".join(u["partners"]) for u in unions
            if any(x not in by_id for x in u["partners"])]
    L.append("ССЫЛКИ В НИКУДА: %d" % len(bad))
    L += ["  " + x for x in bad]
    L.append("")

    L.append("ПАРЫ: %d" % len(unions))
    for u in unions:
        who = " + ".join(by_id[x]["name"] if x in by_id else "?" + x
                         for x in u["partners"])
        mark = " — " + u["src"] if u.get("src") else ""
        if u.get("status") == "divorced":
            mark = " (в разводе)" + mark
        L.append("  %s%s" % (who, mark))
    L.append("")

    L.append("ПРИМЕЧАНИЯ ИЗ ЯЧЕЕК")
    for p in roster:
        if p.get("note"):
            L.append("  %-16s %s" % (p["name"], p["note"]))
    L.append("")

    L.append("ВСЕ ЛЮДИ")
    for p in sorted(roster, key=lambda x: (x.get("gen") or 99, x["name"])):
        L.append("  %s %-18s %-16s %-4s %-6s %-18s %s" % (
            ("%2d" % p["gen"]) if p.get("gen") else "??",
            p["id"], p["name"], p.get("sex") or "?", p.get("line") or "?",
            p.get("parent") or "—", p.get("src", "")))

    REPORT.write_text("\n".join(L) + "\n", "utf-8")


def write_js(roster, unions):
    OUT_JS.parent.mkdir(exist_ok=True)
    rows = []
    for p in sorted(roster, key=lambda x: (x.get("gen") or 99, x.get("born") or 0)):
        rec = {"id": p["id"], "gen": p.get("gen"), "name": p["name"],
               "sex": p.get("sex"), "line": p.get("line")}
        for k in ("alt", "born", "birthday", "died", "parent", "ru", "note"):
            if p.get(k):
                rec[k] = p[k]
        rows.append("    " + json.dumps(rec, ensure_ascii=False))
    pairs = ",\n".join("    " + json.dumps(u, ensure_ascii=False)
                       for u in unions)
    OUT_JS.write_text(
        "/* Собрано из source/АйдарЖапан ШЕЖІРЕСІ-2.xlsx.\n"
        "   Не править руками: правки — в tools/manual.json,\n"
        "   потом python tools/import_xlsx.py */\n\n"
        "window.AIDARZHAPAN = {\n"
        "  /* line: main — прямая линия, rod — род, zhien — жиен,\n"
        "     out — за пределами рода, in — вошедшие в род жёны */\n"
        "  people: [\n" + ",\n".join(rows) + "\n  ],\n\n"
        "  /* Известные супружеские пары. Шежіре — мужская линия,\n"
        "     поэтому жён мало и все они названы поимённо. */\n"
        "  unions: [\n" + pairs + "\n  ]\n};\n", "utf-8")


if __name__ == "__main__":
    main()
